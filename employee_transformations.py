import pandas as pd
import random
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from pathlib import Path

# Role categories for mapping
ROLE_CATEGORIES = {
    'analyst': 'Data & Analysis',
    'engineer': 'Engineering',
    'developer': 'Engineering',
    'manager': 'Management',
    'administrator': 'IT Support',
    'representative': 'Customer Support',
    'vp': 'Executive',
    'nurse': 'Healthcare'
}

# Helper for role category
def map_role_category(title):
    title = str(title).lower()
    for keyword, category in ROLE_CATEGORIES.items():
        if keyword in title:
            return category
    return "Other"

# Helper to generate time
def random_time():
    hour = random.randint(8, 18)  # workday hours
    minute = random.randint(0, 59)
    second = random.randint(0, 59)
    return f"{hour:02}:{minute:02}:{second:02}"

# Helper to generate day name
def random_day():
    return random.choice(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])

# Main transformation function
def transform_employee_data(input_path, output_path):
    df = pd.read_csv(input_path)

    # Full name
    df['full_name'] = df['first_name'].astype(str) + ' ' + df['last_name'].astype(str)

    # Convert 'date' to datetime, coerce errors
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    now = pd.Timestamp.now()

    # Handle day, time, and working_days
    df['day'] = df['date'].dt.day_name()
    df['time'] = df['date'].dt.strftime('%H:%M:%S')
    df['working_days'] = (now - df['date']).dt.days
    df['day'] = df['date'].apply(lambda x: x.day_name() if pd.notnull(x) else random_day())
    df['time'] = df['date'].apply(lambda x: x.strftime('%H:%M:%S') if pd.notnull(x) and x.time() != datetime.min.time() else random_time())
    df['working_days'] = df['date'].apply(lambda x: (now - x).days if pd.notnull(x) else 0)


    # Job title handling
    df['job_title'] = df['job_title'].astype(str)
    df['job_prefix'] = df['job_title'].str[:3]
    df['job_suffix'] = df['job_title'].str[-3:]

    # Duplicate detection
    df['is_duplicate_role'] = df.duplicated(['job_title'], keep=False)
    df['dup_count'] = df.groupby('job_title').cumcount()
    df['unique_role_key'] = df.apply(
        lambda x: f"{x['job_title'].lower()}_{x['dup_count']}" if x['is_duplicate_role'] else x['job_title'].lower(),
        axis=1
    )

    # Frequency
    df['job_title_frequency'] = df.groupby('job_title')['job_title'].transform('count')

    # Map role category
    df['role_category'] = df['job_title'].apply(map_role_category)

    # Final output dataframe
    df_final = df[['full_name', 'day', 'time', 'job_title', 'job_prefix', 'job_suffix',
                   'is_duplicate_role', 'unique_role_key', 'working_days',
                   'job_title_frequency', 'role_category']]

    # Save with context manager
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False)

    # Post-formatting the Excel file
    format_excel(output_path)

# Formatting Excel output
def format_excel(output_path):
    wb = load_workbook(output_path)
    ws = wb.active

    # Center alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Enable gridlines (default behavior)
    ws.sheet_view.showGridLines = True

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save formatted Excel
    wb.save(output_path)

# Run the function on your uploaded file
input_csv = "Employee_records.csv"
output_excel = "Advanced_Employee_Report.xlsx"
transform_employee_data(input_csv, output_excel)
